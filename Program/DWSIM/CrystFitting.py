#%%%%%%%%%% New version of the Nelder-Mead Simplex script %%%%%%%%%%#

#Filepath for input and output worksheets/JSON files

import numpy as np
import random
import openpyxl
import sys
import subprocess
import os
import json
from datetime import datetime
from statistics import mean
import os, sys

#SystemPath = "C:\\Users\\Maxim\\Desktop\\Input\\"
#SystemPath = os.getcwd() + "\\Program\\DWSIM\\"
SystemPath = os.path.join(os.path.dirname(__file__))

#Get Inputs from subprocess:
ExDataPath = str(sys.argv[1]) #The filepath is the first argument passed via the subprocess.Popen method
Iterations = int(float(sys.argv[2])) #Number of iterations. Can be set via the box in the GUI


#Variables for debugging purposes
#ExDataPath = "C:\\Users\\Maxim\\Desktop\\Input\\ExperimentalData_Prog4.xlsx"
#Iterations = 2

startTime = datetime.now()
with open(SystemPath + "\\Output\\ParaFitCryst_CalcuLog.txt", "w") as myfile:
    myfile.write("Simulation number, Time, Type, Objective function, Growth Parameter 1, Growth Parameter 2, Growth Parameter 3, Agglomeration Parameter 1, Agglomeration Parameter 2, Birth Parameter 1, Birth Parameter 2, Birth Parameter 3\n")
with open(SystemPath + "\\Output\\ParaFitCryst_ObjectiveLog.txt", "w") as myfile:
    myfile.write("Simulation number, Time, Objective function, Mean Objective, Growth Parameter 1, Growth Parameter 2, Growth Parameter 3, Agglomeration Parameter 1, Agglomeration Parameter 2, Birth Parameter 1, Birth Parameter 2, Birth Parameter 3\n")

#Get experimental data
ExData = openpyxl.load_workbook(ExDataPath)
ExDataSheet = ExData['Results']
Ex10, Ex25, Ex50, Ex75, Ex90 = np.zeros(4), np.zeros(4), np.zeros(4), np.zeros(4), np.zeros(4)

for i in range(4):
    Ex10[i] = ExDataSheet.cell(row=i+2, column=2).value
    Ex25[i] = ExDataSheet.cell(row=i+2, column=3).value
    Ex50[i] = ExDataSheet.cell(row=i+2, column=4).value
    Ex75[i] = ExDataSheet.cell(row=i+2, column=5).value
    Ex90[i] = ExDataSheet.cell(row=i+2, column=6).value
ExData.close()
#SimResults stores the calculated diameters from the simulation for the respective time
SimResults10, SimResults25, SimResults50, SimResults75, SimResults90 = np.zeros(4), np.zeros(4), np.zeros(4), np.zeros(4), np.zeros(4)
#CSD is used to store the particle size distributions obtained in the simulations
CSD30, CSD60, CSD90, CSD120 = np.zeros(30), np.zeros(30), np.zeros(30), np.zeros(30)

#Initialization of the Simplex
#Each entry in the dictionary stores a value used in the initial simplex calculation
Simplex = {
    1: np.zeros(5),
    2: np.zeros(5),
    3: np.zeros(5),
    4: np.zeros(5),
    5: np.zeros(5),
    6: np.zeros(5)
}

def Load_JSON_file(filepath):
    """
    Loads the .json file used for storing the input parameters in the DWSIM simulation.
    The data is stored as dictionary and returned by this function
    """
    JSON_file = open(filepath + "\\Input.json")
    JSON_data = json.load(JSON_file)
    return JSON_data

#Initialize parameters used in the calculation
Parameters = Load_JSON_file(SystemPath)
Time = int(int(Parameters["CycleTime"]) * int(Parameters["CrystallizationModules"]))
Toleranz = 5
delta = int(Parameters["Delta"])
Number = 430000 #Total number of particles. Obtained using the "Estimations.xlsx" worksheet #Not needed???

#Initialize lists used in the calculation
PenaltySimp = np.zeros(6) #Used to store values of the objective function for the simplex
ResCSD = np.zeros(30)


#Import from the opened JSON file
Widths = Parameters["ClassWidth"]
Length = Parameters["Length"]
GrowthRateApproach = Parameters["GrowthRate"]

#Analysis of simulation results
#Results are imported from Excel and d01, d10, d50, d90 and d99 are determined and returned
def ImportSimulationOutput():
    """
    Loads the results of the simulation, converts it from a density function to a sum function and calculates dxx.
    The parameters are returned in a list for further use in calculations
    """

    SimOutput = openpyxl.load_workbook(SystemPath + '\\Output\\CrystallizationOutput.xlsx')
    SimOut = SimOutput['Results']
    for Ex in range(30):
        CSD30[Ex] = SimOut.cell(row=Ex+9, column=(1+(30*60)//delta)).value
        CSD60[Ex] = SimOut.cell(row=Ex+9, column=(1+(60*60)//delta)).value
        CSD90[Ex] = SimOut.cell(row=Ex+9, column=(1+(90*60)//delta)).value
        CSD120[Ex] = SimOut.cell(row=Ex+9, column=(1+(120*60)//delta)).value
    Distros = [CSD30, CSD60, CSD90, CSD120]

    #Auxiliary lists for the calculation:
    NumDest, Q3 = np.zeros(30), np.zeros(30)

    for j in range(4):
        NumDest_tot = 0
        #for i in range(0, 30):
        #    #Conversion of the density function to a cumulative function
        #    ResCSD[i] = Distros[j][i] * Widths[i]
        #    if i == 0:
        #        ResCSD[i] = ResCSD[i]
        #    else:
        #        ResCSD[i] = ResCSD[i] + ResCSD[i-1]
        #for i in range(0, 30):
        #    #Normalization of the distribution function
        #    #Done to mitigate numerical instabilities and numerical diffusion
        #    ResCSD[i] = ResCSD[i] / ResCSD[29]
        #
        #    if i == 0:
        #        NumDest[i] = ResCSD[i]
        #    else:
        #        NumDest[i] = (ResCSD[i] - ResCSD[i-1])
        #    NumDest_i = NumDest[i] * Length[i]**3
        #    NumDest_tot = NumDest_tot + NumDest_i
        #for i in range(0, 30):
        #    #Conversion of the number density distribution to a Q3 distribution
        #    Q3[i] = NumDest[i] * Length[i]**3 / NumDest_tot
        #    if i == 0:
        #        Q3[i] = Q3[i]
        #    else:
        #        Q3[i] = Q3[i] + Q3[i-1]

        #Generate the Q3 distribution from the output DELTA Q3 distribution from the Excel Sheet
        for i in range(0, 30):
            if i == 0:
                Q3[i] = Distros[j][i]
            else:
                Q3[i] = Distros[j][i] + Q3[i-1]


        for i in range(1, 30):
            #
            if Q3[i] < 0.01:
                i+=1
            #if (Q3[i] >= 0.01 and Q3[i-1] < 0.01).all():
            if np.all(np.logical_and(Q3[i] >= 0.10, Q3[i-1] < 0.10)):
                SimResults10[j] = (Length[i] - Length[i-1])/(Q3[i] - Q3[i-1]) * (0.10 - Q3[i-1]) + Length[i-1]
            if np.all(np.logical_and(Q3[i] >= 0.25, Q3[i-1] < 0.25)):
                SimResults25[j] = (Length[i] - Length[i-1])/(Q3[i] - Q3[i-1]) * (0.25 - Q3[i-1]) + Length[i-1]
            if np.all(np.logical_and(Q3[i] >= 0.50, Q3[i-1] < 0.50)):
                SimResults50[j] = (Length[i] - Length[i-1])/(Q3[i] - Q3[i-1]) * (0.50 - Q3[i-1]) + Length[i-1]
            if np.all(np.logical_and(Q3[i] >= 0.75, Q3[i-1] < 0.75)):
                SimResults75[j] = (Length[i] - Length[i-1])/(Q3[i] - Q3[i-1]) * (0.75 - Q3[i-1]) + Length[i-1]
            if np.all(np.logical_and(Q3[i] >= 0.90, Q3[i-1] < 0.90)):
                SimResults90[j] = (Length[i] - Length[i-1])/(Q3[i] - Q3[i-1]) * (0.90 - Q3[i-1]) + Length[i-1]
                break
    SimOutput.close()
    os.remove(SystemPath + '\\Output\\CrystallizationOutput.xlsx')
    return [SimResults10, SimResults25, SimResults50, SimResults75, SimResults90]

#Populate the initial simplex with values:
if GrowthRateApproach == "Exponential":
    for i in range(1,7):
        Simplex[i][0] = random.uniform(1000000, 10000000) #Prefactor
        Simplex[i][1] = random.uniform(50000, 75000) #Activation energy
        Simplex[i][2] = random.uniform(1, 2) #Exponential
        Simplex[i][3] = random.uniform(20, 250) #Critical length
        Simplex[i][4] = random.uniform(50, 1000) #Agglomeration parameter

else:
    for i in range(1,7):
        Simplex[i][0] = random.uniform(0.1, 100) #A_BCF
        Simplex[i][1] = random.uniform(0.01, 0.5) #B_BCF
        Simplex[i][2] = random.uniform(0, 0) #Unused
        Simplex[i][3] = random.uniform(20, 250) #Critical length
        Simplex[i][4] = random.uniform(50, 1000) #Agglomeration parameter


#Simplex[1][0], Simplex[1][1], Simplex[1][2], Simplex[1][3], Simplex[1][4] = 0.3, 4700, 1.089, 193, 0.0169
#Simplex[2][0], Simplex[2][1], Simplex[2][2], Simplex[2][3], Simplex[2][4] = 7.77, 7195, 1.815, 127.84, 0.273
#Simplex[3][0], Simplex[3][1], Simplex[3][2], Simplex[3][3], Simplex[3][4] = 7.24, 6901, 1.477, 75.1, 0.0124
#Simplex[4][0], Simplex[4][1], Simplex[4][2], Simplex[4][3], Simplex[4][4] = 3.429, 7291, 1.32, 135.755, 0.0304
#Simplex[5][0], Simplex[5][1], Simplex[5][2], Simplex[5][3], Simplex[5][4] = 5.93, 6974, 1.03, 70.977, 0.193
#Simplex[6][0], Simplex[6][1], Simplex[6][2], Simplex[6][3], Simplex[6][4] = 0.7195, 9170, 1.89, 199, 0.4015

def PenaltyCalculation(lstSimulationOutput):
    """
    Calculation of the penalty function which is minimized over the course of this script.
    The input for this function is a list containing lists containing the time dependent results for
    d01, d10, d50, d90 and d99
    """
    Objective = 0
    #Objective_i_10 = ((lstSimulationOutput[0][0] - Ex10[0])**2 + (lstSimulationOutput[0][1]-Ex10[1])**2 + (lstSimulationOutput[0][2]-Ex10[2])**2 +(lstSimulationOutput[0][3]-Ex10[3])**2)
    #Objective_i_25 = ((lstSimulationOutput[1][0] - Ex25[0])**2 + (lstSimulationOutput[1][1]-Ex25[1])**2 + (lstSimulationOutput[1][2]-Ex25[2])**2 +(lstSimulationOutput[1][3]-Ex25[3])**2)
    #Objective_i_50 = ((lstSimulationOutput[2][0] - Ex50[0])**2 + (lstSimulationOutput[2][1]-Ex50[1])**2 + (lstSimulationOutput[2][2]-Ex50[2])**2 +(lstSimulationOutput[2][3]-Ex50[3])**2)
    #Objective_i_75 = ((lstSimulationOutput[3][0] - Ex75[0])**2 + (lstSimulationOutput[3][1]-Ex75[1])**2 + (lstSimulationOutput[3][2]-Ex75[2])**2 +(lstSimulationOutput[3][3]-Ex75[3])**2)
    #Objective_i_90 = ((lstSimulationOutput[4][0] - Ex90[0])**2 + (lstSimulationOutput[4][1]-Ex90[1])**2 + (lstSimulationOutput[4][2]-Ex90[2])**2 +(lstSimulationOutput[4][3]-Ex90[3])**2)
    #Objective_i_Width = ((lstSimulationOutput[4][0] - lstSimulationOutput[0][0]) - (Ex90[0] - Ex10[0]))**2 + ((lstSimulationOutput[4][1] - lstSimulationOutput[0][1]) - (Ex90[1] - Ex10[1]))**2 + ((lstSimulationOutput[4][2] - lstSimulationOutput[0][2]) - (Ex90[2] - Ex10[2]))**2 + ((lstSimulationOutput[4][3] - lstSimulationOutput[0][3]) - (Ex90[3] - Ex10[3]))**2
    
    Objective_i_10 = (lstSimulationOutput[0][3]-Ex10[3])**2
    Objective_i_25 = (lstSimulationOutput[1][3]-Ex25[3])**2
    Objective_i_50 = (lstSimulationOutput[2][3]-Ex50[3])**2
    Objective_i_75 = (lstSimulationOutput[3][3]-Ex75[3])**2
    Objective_i_90 = (lstSimulationOutput[4][3]-Ex90[3])**2
    Objective_i_Width = ((lstSimulationOutput[4][3] - lstSimulationOutput[0][3]) - (Ex90[3] - Ex10[3]))**2

    #Objective = Objective_i_10 + Objective_i_25 + Objective_i_50 + Objective_i_75 + Objective_i_90 + Objective_i_Width
    #Objective = Objective_i_10 + Objective_i_50 + Objective_i_90 + Objective_i_Width
    Objective = Objective_i_50 + Objective_i_25 + Objective_i_75
    return np.sqrt(Objective)

def AverageSimplex():
    """
    Averages the input values stored in the simplex dictionary.
    Used in the calculation of the reflection, expansion and contraction step.
    """
    Ave = np.empty(5)
    for value in range(5):
        Ave[value] = (Simplex[1][value] + Simplex[2][value] + Simplex[3][value] + Simplex[4][value] + Simplex[5][value] + Simplex[6][value])/6
    return Ave

def Calculate(intSimNo):
    """
    This function calls the DWSIM solver using the "subprocess.call" method.
    The solver is called via an additional script to prevent a synchrounous calculation of this script and the DWSIM simulation which would cause erroneous results.
    """
    Process = subprocess.Popen(['python',
                                SystemPath + '\\RunCrystallization.py',
                                SystemPath + "\\DWSIM_Files\\",
                                "Fitting"], text=True)
    Process.wait()

def GiveInputs(lstIn):
    """
    Passes arguments for growth and agglomeration parameters to the .JSON file which is subsequently read by the DWSIM script.
    """
    JSONFile = Load_JSON_file(SystemPath)
    JSONFile["GrowthConstant1"] = lstIn[0]
    JSONFile["GrowthConstant2"] = lstIn[1]
    JSONFile["GrowthConstant3"] = lstIn[2]
    JSONFile["AgglConstant1"] = lstIn[3]
    JSONFile["AgglConstant2"] = lstIn[4]
    with open(SystemPath + "\\Input.json", "w") as json_file:
        json.dump(JSONFile, json_file, indent=4)

def ReplaceSimplex(arrVals, SimpNo):
    """
    This function is called if either the reflection, expansion or contraction step prove to be successful.
    By replacing the values in the simplex, the original one can be kept throughout the script instead of being calculated again for each iteration.
    """
    for Repl in range(5):
        Simplex[SimpNo][Repl] = arrVals[Repl]

def ReplaceMin(fltVal, MinNo):
    PenaltySimplex[MinNo] = fltVal

def Reflection(lstAve, intNumber):
    """
    Application of the reflection step in the Nelder-mead method.
    Initial step performed after the initial simplex calculation.
    """    
    alpha = 1 #Parameter defining the step width in the reflection calculation
    Reflect = abs((1 + alpha) * lstAve - alpha * Simplex[list(PenaltySimplex).index(max(PenaltySimplex)) + 1])
    GiveInputs(Reflect)
    Calculate(intNumber)
    PenaltyReflection = PenaltyCalculation(ImportSimulationOutput())
    #print(PenaltyReflection)
    with open(SystemPath + "\\Output\\ParaFitCryst_CalcuLog.txt", "a") as myfile:
        myfile.write("{}, {}, {}, {}, {}, {}, {}, {}, {}\n".format(intNumber,
                                                                   datetime.now().strftime("%H:%M:%S"),
                                                                   "Reflection",
                                                                   PenaltyReflection,
                                                                   Reflect[0],
                                                                   Reflect[1],
                                                                   Reflect[2],
                                                                   Reflect[3],
                                                                   Reflect[4]))
    return PenaltyReflection, Reflect

def Expansion(ReflecPoint, lstAve, intNumber):
    """
    Performs the expansion step in the Nelder-Mead method.
    The absolute value is used in the determination of input parameters in order to prevent negative values.
    Should negative values occur, the simulation in DWSIM cannot be performed accordingly.
    """
    gamma = 0.5 #Parameter used to define the extension of the simplex in the expansion step
    Expand = abs((1 + gamma) * ReflecPoint - gamma * lstAve)
    GiveInputs(Expand)
    Calculate(intNumber)
    PenaltyExpansion = PenaltyCalculation(ImportSimulationOutput())
    #print(PenaltyExpansion)
    with open(SystemPath + "\\Output\\ParaFitCryst_CalcuLog.txt", "a") as myfile:
        myfile.write("{}, {}, {}, {}, {}, {}, {}, {}, {}\n".format(intNumber,
                                                                   datetime.now().strftime("%H:%M:%S"),
                                                                   "Expansion",
                                                                   PenaltyExpansion,
                                                                   Expand[0],
                                                                   Expand[1],
                                                                   Expand[2],
                                                                   Expand[3],
                                                                   Expand[4]))
    return PenaltyExpansion, Expand

def Contraction(lstAve, intNumber):
    """
    Performs the contraction step in the Nelder Mead method.
    """
    beta = 0.5 #Parameter defining the extent of the contraction
    Contract = abs(beta * Simplex[list(PenaltySimplex).index(min(PenaltySimplex)) + 1] + (1 - beta) * lstAve)
    GiveInputs(Contract)
    Calculate(intNumber)
    PenaltyContraction = PenaltyCalculation(ImportSimulationOutput())
    #print(PenaltyContraction)
    with open(SystemPath + "\\Output\\ParaFitCryst_CalcuLog.txt", "a") as myfile:
        myfile.write("{}, {}, {}, {}, {}, {}, {}, {}, {}\n".format(intNumber,
                                                                   datetime.now().strftime("%H:%M:%S"),
                                                                   "Contraction",
                                                                    PenaltyContraction,
                                                                    Contract[0],
                                                                    Contract[1],
                                                                    Contract[2],
                                                                    Contract[3],
                                                                    Contract[4]))
    return PenaltyContraction, Contract

#Calculation of the initial Simplex:
SimulationNo = 1
Average = AverageSimplex()
Counter = 0
#print(Simplex, Average)

PenaltySimplex = [None] * 6
for i in range(0, 6):
    GiveInputs(Simplex[i + 1])
    Calculate(SimulationNo)
    PenaltySimplex[i] = PenaltyCalculation(ImportSimulationOutput())
    #print(PenaltySimplex)
    with open(SystemPath + "\\Output\\ParaFitCryst_CalcuLog.txt", "a") as myfile:
        myfile.write("{}, {}, {}, {}, {}, {}, {}, {}, {}\n".format(SimulationNo,
                                                                   datetime.now().strftime("%H:%M:%S"),
                                                                   "Initial Simplex",
                                                                   PenaltySimplex[i],
                                                                   Simplex[i + 1][0],
                                                                   Simplex[i + 1][1],
                                                                   Simplex[i + 1][2],
                                                                   Simplex[i + 1][3],
                                                                   Simplex[i + 1][4]))

if min(PenaltySimplex) <= Toleranz:
    print([1, SimulationNo,
           min(PenaltySimplex),
           mean(PenaltySimplex),
           Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][0],
           Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][1],
           Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][2],
           Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][3],
           Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][4]])
    sys.stdout.flush()
    sys.exit()
else:
    while SimulationNo <= Iterations: #Maximum number of iterations
        Average = AverageSimplex()
        ReflectionResults = Reflection(Average, SimulationNo)

        if ReflectionResults[0] <= Toleranz:
            #print("Final simplex:", Simplex[list(PenaltySimplex).index(min(PenaltySimplex)) + 1])
            #print("Final value of optimization function:", min(PenaltySimplex))
            print([1, SimulationNo,
                   min(PenaltySimplex),
                   mean(PenaltySimplex),
                   Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][0],
                   Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][1],
                   Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][2],
                   Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][3],
                   Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][4]])
            sys.stdout.flush()
            sys.exit()
        else: 
            if ReflectionResults[0] < min(PenaltySimplex):
                ExpansionResults = Expansion(ReflectionResults[1], Average, SimulationNo)

                if ExpansionResults[0] <= Toleranz:
                    #print("Optimal value:", ExpansionResults[1])
                    #print("Final value of optimization function:", ExpansionResults[0])
                    print([1, SimulationNo,
                           min(PenaltySimplex),
                           mean(PenaltySimplex),
                           Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][0],
                           Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][1],
                           Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][2],
                           Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][3],
                           Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][4]])
                    sys.stdout.flush()
                    sys.exit()

                else:
                    if ExpansionResults[0] < min(PenaltySimplex):
                        ReplaceSimplex(ExpansionResults[1], list(PenaltySimplex).index(max(PenaltySimplex))+1)
                        ReplaceMin(ExpansionResults[0], list(PenaltySimplex).index(max(PenaltySimplex)))
                        #print(PenaltySimplex)
                    else:
                        ReplaceSimplex(ReflectionResults[1], list(PenaltySimplex).index(max(PenaltySimplex))+1)
                        ReplaceMin(ReflectionResults[0], list(PenaltySimplex).index(max(PenaltySimplex)))
                        #print(PenaltySimplex)
            ##### Middle branch of the Nelder Mead algorithm #####
            elif any(i > ReflectionResults[0] for i in PenaltySimplex if i != max(PenaltySimplex)):
                ReplaceSimplex(ReflectionResults[1], list(PenaltySimplex).index(max(PenaltySimplex))+1)
                ReplaceMin(ReflectionResults[0], list(PenaltySimplex).index(max(PenaltySimplex)))
                #print(PenaltySimplex)
            
            ##### Right branch of the Nelder Mead algorithm #####
            else:

                if ReflectionResults[0] < max(PenaltySimplex):
                    ReplaceSimplex(ReflectionResults[1], list(PenaltySimplex).index(max(PenaltySimplex))+1)
                    ReplaceMin(ReflectionResults[0], list(PenaltySimplex).index(max(PenaltySimplex)))
                    Average = AverageSimplex()
                    #print(PenaltySimplex)
                
                ##### Contraction step #####
                ContractionResults = Contraction(Average, SimulationNo)

                if ContractionResults[0] <= Toleranz:
                    #print("Optimal value:", ContractionResults[1])
                    #print("Final value of optimization function:", ContractionResults[0])
                    print([1, SimulationNo,
                           min(PenaltySimplex),
                           mean(PenaltySimplex),
                           Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][0],
                           Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][1],
                           Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][2],
                           Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][3],
                           Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][4]])
                    sys.stdout.flush()
                    sys.exit()

                elif ContractionResults[0] < max(PenaltySimplex):
                    ReplaceSimplex(ContractionResults[1], list(PenaltySimplex).index(max(PenaltySimplex))+1)
                    ReplaceMin(ContractionResults[0], list(PenaltySimplex).index(max(PenaltySimplex)))
                    #print(PenaltySimplex)

                else:
                    for i in range(1,7):
                        Simplex[i] = abs((Simplex[i]+ Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1])/2)
                    Average = AverageSimplex()
                    #print("Calculating compressed simplex...")
                    for i in range(1,7):
                        GiveInputs(Simplex[i])
                        Calculate(SimulationNo)
                        PenaltySimplex[i-1] = PenaltyCalculation(ImportSimulationOutput())
                        with open(SystemPath + "\\Output\\ParaFitCryst_CalcuLog.txt", "a") as myfile:
                            myfile.write("{}, {}, {}, {}, {}, {}, {}, {}, {}\n".format(SimulationNo,
                                                                                       datetime.now().strftime("%H:%M:%S"),
                                                                                       "Compression",
                                                                                       PenaltySimplex[i-1],
                                                                                       Simplex[i][0],
                                                                                       Simplex[i][1],
                                                                                       Simplex[i][2],
                                                                                       Simplex[i][3],
                                                                                       Simplex[i][4]))
                    #print("Finished compressed simplex calculation.")
                    #print(PenaltySimplex)
                    if min(PenaltySimplex) <= Toleranz:
                        #print("Optimal value:", Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1])
                        #print("Final value of optimization function: ", min(PenaltySimplex))
                        print([1, SimulationNo,
                               min(PenaltySimplex),
                               mean(PenaltySimplex),
                               Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][0],
                               Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][1],
                               Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][2],
                               Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][3],
                               Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][4]])
                        sys.stdout.flush()
                        sys.exit()
        #print("Minimum of the objective function in iteration %i was:" %SimulationNo, min(PenaltySimplex))
        with open(SystemPath + "\\Output\\ParaFitCryst_ObjectiveLog.txt", "a") as myfile:
            myfile.write("{}, {}, {}, {}, {}, {}, {}, {}, {}\n".format(SimulationNo,
                                                                       datetime.now().strftime("%H:%M:%S"),
                                                                       min(PenaltySimplex),
                                                                       mean(PenaltySimplex),
                                                                       Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][0],
                                                                       Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][1],
                                                                       Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][2],
                                                                       Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][3],
                                                                       Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][4]))
        if SimulationNo != Iterations:
            print([0, SimulationNo,
                   min(PenaltySimplex),
                   mean(PenaltySimplex),
                   Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][0],
                   Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][1],
                   Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][2],
                   Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][3],
                   Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][4]])
            sys.stdout.flush()

        if mean(PenaltySimplex)/min(PenaltySimplex) <= 1.025:
            Counter = Counter + 1
        if Counter >= 15:
            break
        SimulationNo += 1


#End the script if convergence is not achieved within the predetermined number of calculation steps:
#print("Simulation did not converge in %i iterations." %SimulationNo)
#print("The best value was achieved for: ", Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1])
#print("Corresponding value of the objective function:", min(PenaltySimplex))

#Pass the best found value combination to the .JSON file and perform one simulation to generate results
GiveInputs(Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1])
Calculate(SimulationNo + 1)

#The first entry provides an indication whether an exit point has been reached. If the value is 1, it indicates to the calling script that the calculation here is finished
#If the exit code is 0, the calling script knows that the calculation is still running
print([1, SimulationNo,
       min(PenaltySimplex),
       mean(PenaltySimplex),
       Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][0],
       Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][1],
       Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][2],
       Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][3],
       Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][4]])
sys.stdout.flush()
