import numpy as np
import openpyxl
import sys
import os
import subprocess
import json
import time
from datetime import datetime
from statistics import mean
import random

SystemPath = os.path.dirname(__file__)

ExDataPath = str(sys.argv[1])
#ExDataPath = "C:\\Users\\Maxim\\Desktop\\Input\\ExperimentalData_Oszi.xlsx"
Iterations = int(float(sys.argv[2]))
#Iterations = 40


startTime = datetime.now()

#Prepare log file
with open(SystemPath + "\\Output\\ParaFitCryst_CalcuLog.txt", "w") as myfile:
    myfile.write("Simulation number, Time, Type, Objective function, Objective function 2, Growth Parameter 1, Growth Parameter 2, Growth Parameter 3, Agglomeration Parameter 1, Agglomeration Parameter 2, Fitting algorithm \n")
with open(SystemPath + "\\Output\\ParaFitCryst_ObjectiveLog.txt", "w") as myfile:
    myfile.write("Simulation number, Time, Objective function, Mean Objective, Objective function 2, Growth Parameter 1, Growth Parameter 2, Growth Parameter 3, Agglomeration Parameter 1, Agglomeration Parameter 2, Fitting algorithm\n")

#Read Experimental Data from input workbook
#The input workbook is passed as input argument by the subprocess
ExData = openpyxl.load_workbook(ExDataPath)
ExDataSheet = ExData['Results']
Ex10, Ex25, Ex50, Ex75, Ex90 = np.zeros(4), np.zeros(4), np.zeros(4), np.zeros(4), np.zeros(4)

for i in range(4):
    Ex10[i] = ExDataSheet.cell(row=i+2, column=2).value
    Ex25[i] = ExDataSheet.cell(row=i+2, column=3).value
    Ex50[i] = ExDataSheet.cell(row=i+2, column=4).value
    Ex75[i] = ExDataSheet.cell(row=i+2, column=5).value
    Ex90[i] = ExDataSheet.cell(row=i+2, column=6).value
ExData.close()

#SimResults stores the calculated diameters from the simulation for the respective time
SimResults10, SimResults25, SimResults50, SimResults75, SimResults90 = np.zeros(4), np.zeros(4), np.zeros(4), np.zeros(4), np.zeros(4)
#CSD is used to store the particle size distributions obtained in the simulations
CSD30, CSD60, CSD90, CSD120 = np.zeros(30), np.zeros(30), np.zeros(30), np.zeros(30)

def Load_JSON_file(filepath):
    """
    Loads the .json file used for storing the input parameters in the DWSIM simulation.
    The data is stored as dictionary and returned by this function
    """
    #JSON_file = open(filepath + "\\DWSIM\\Input.json")
    JSON_file = open(filepath + "\\Input.json")
    JSON_data = json.load(JSON_file)
    return JSON_data

#Initialize parameters used in the calculation
Parameters = Load_JSON_file(SystemPath)
Time = int(int(Parameters["CycleTime"]) * int(Parameters["CrystallizationModules"]))
Toleranz = 5
delta = int(Parameters["Delta"])

GrowthRateApproach = Parameters["GrowthRate"]

#Initialize lists used in the calculation
PenaltySimp = np.zeros(6) #Used to store values of the objective function for the simplex
ResCSD = np.zeros(30)


#Import from the opened JSON file
Widths = Parameters["ClassWidth"]
Length = Parameters["Length"]

ShapeFactor = Parameters["ShapeFactor"]
rho_Cryst = Parameters["CrystalDensity"]

def ImportSimulationOutput():
    """
    Loads the results of the simulation, converts it from a density function to a sum function and calculates dxx.
    The parameters are returned in a list for further use in calculations
    """

    SimOutput = openpyxl.load_workbook(SystemPath + '\\Output\\CrystallizationOutput.xlsx')
    SimOut = SimOutput['Results']
    MassFinal = SimOut.cell(row=5, column=(1+(120*60)//delta)).value
    for Ex in range(30):
        CSD30[Ex] = SimOut.cell(row=Ex+9, column=(1+(30*60)//delta)).value
        CSD60[Ex] = SimOut.cell(row=Ex+9, column=(1+(60*60)//delta)).value
        CSD90[Ex] = SimOut.cell(row=Ex+9, column=(1+(90*60)//delta)).value
        CSD120[Ex] = SimOut.cell(row=Ex+9, column=(1+(120*60)//delta)).value
    Distros = [CSD30, CSD60, CSD90, CSD120]

    #Auxiliary lists for the calculation:
    NumDest, Q3 = np.zeros(30), np.zeros(30)
    num = 0
    for n in range(30):
        n_i = Distros[3][n] * MassFinal / (ShapeFactor * rho_Cryst * (Length[n]*10**(-6))**3)
        num = num + n_i
    Numero = num
    MassSimOut = SimOut.cell(row=5, column=(1+(120*60)/delta)).value * 10**3


    for j in range(4):
        NumDest_tot = 0
        
        for i in range(0, 30):
            if i == 0:
                Q3[i] = Distros[j][i]
            else:
                Q3[i] = Distros[j][i] + Q3[i-1]
        for i in range(1, 30):
            #
            if Q3[i] < 0.01:
                i+=1
            #if (Q3[i] >= 0.01 and Q3[i-1] < 0.01).all():
            if np.all(np.logical_and(Q3[i] >= 0.10, Q3[i-1] < 0.10)):
                SimResults10[j] = (Length[i] - Length[i-1])/(Q3[i] - Q3[i-1]) * (0.10 - Q3[i-1]) + Length[i-1]
            if np.all(np.logical_and(Q3[i] >= 0.25, Q3[i-1] < 0.25)):
                SimResults25[j] = (Length[i] - Length[i-1])/(Q3[i] - Q3[i-1]) * (0.25 - Q3[i-1]) + Length[i-1]
            if np.all(np.logical_and(Q3[i] >= 0.50, Q3[i-1] < 0.50)):
                SimResults50[j] = (Length[i] - Length[i-1])/(Q3[i] - Q3[i-1]) * (0.50 - Q3[i-1]) + Length[i-1]
            if np.all(np.logical_and(Q3[i] >= 0.75, Q3[i-1] < 0.75)):
                SimResults75[j] = (Length[i] - Length[i-1])/(Q3[i] - Q3[i-1]) * (0.75 - Q3[i-1]) + Length[i-1]
            if np.all(np.logical_and(Q3[i] >= 0.90, Q3[i-1] < 0.90)):
                SimResults90[j] = (Length[i] - Length[i-1])/(Q3[i] - Q3[i-1]) * (0.90 - Q3[i-1]) + Length[i-1]
                break
    SimOutput.close()
    os.remove(SystemPath + '\\Output\\CrystallizationOutput.xlsx')
    return [SimResults10, SimResults25, SimResults50, SimResults75, SimResults90, Numero]

def PenaltyCalculation(lstSimulationOutput):
    """
    Calculation of the penalty function which is minimized over the course of this script.
    The input for this function is a list containing lists containing the time dependent results for
    d01, d10, d50, d90 and d99
    """
    
    Objective_i_10 = (lstSimulationOutput[0][3]-Ex10[3])**2
    Objective_i_25 = (lstSimulationOutput[1][3]-Ex25[3])**2
    Objective_i_50 = (lstSimulationOutput[2][3]-Ex50[3])**2
    Objective_i_75 = (lstSimulationOutput[3][3]-Ex75[3])**2
    Objective_i_90 = (lstSimulationOutput[4][3]-Ex90[3])**2
    Objective_i_Width = ((lstSimulationOutput[4][3] - lstSimulationOutput[0][3]) - (Ex90[3] - Ex10[3]))**2

    #Objective = Objective_i_10 + Objective_i_25 + Objective_i_50 + Objective_i_75 + Objective_i_90 + Objective_i_Width
    Objective = Objective_i_50
    Objective_Number = np.sqrt((lstSimulationOutput[5]-2246753.358)**2)
    Objective3 = Objective_i_Width
    return [np.sqrt(Objective + Objective_i_25 + Objective_i_75), Objective_Number, np.sqrt(Objective3)]
    
def Calculate():
    """
    This function calls the DWSIM solver using the "subprocess.call" method.
    The solver is called via an additional script to prevent a synchrounous calculation of this script and the DWSIM simulation which would cause erroneous results.
    """
    Process = subprocess.Popen(['python', SystemPath + '\\RunCrystallization.py', SystemPath + "\\DWSIM_Files\\", "Fitting"], text=True)
    Process.wait()

def GiveInputs(lstIn):
    """
    Passes arguments for growth and agglomeration parameters to the .JSON file which is subsequently read by the DWSIM script.
    """
    JSONFile = Load_JSON_file(SystemPath)
    JSONFile["GrowthConstant1"] = lstIn[0]
    JSONFile["GrowthConstant2"] = lstIn[1]
    JSONFile["GrowthConstant3"] = lstIn[2]
    JSONFile["AgglConstant1"] = lstIn[3]
    JSONFile["AgglConstant2"] = lstIn[4]
    JSONFile["BirthConstant1"] = lstIn[5]
    JSONFile["BirthConstant2"] = lstIn[6]
    JSONFile["BirthConstant3"] = lstIn[7]
    with open(SystemPath + "\\Input.json", "w") as json_file:
        json.dump(JSONFile, json_file, indent=4)

####################### Functions specific for PSO #######################
def ObjectiveFunction(para1:float, para2:float, para3:float, para4:float, para5:float, para6:float, para7:float, para8:float):

    GiveInputs([para1, para2, para3, para4, para5, para6, para7, para8])
    Calculate()
    Output = PenaltyCalculation(ImportSimulationOutput())
    return Output
    
def update(SimuNo:int):
    global V, X, pbest, pbest_obj, gbest, gbest_obj, V2, X2, pbest2, pbest_obj2, gbest2, gbest_obj2

    r1, r2 = np.random.rand(2)
    V = w * V + c1 * r1 *(pbest - X) + c2 * r2 * (gbest.reshape(-1,1) - X)
    V2 = w * V2 + c1 * r1 *(pbest2 - X2) + c2 * r2 * (gbest2.reshape(-1,1) - X2)
    X = X + V
    X2 = X2 + V2
    
    obj = np.empty(n_particles)
    obj2 = np.empty(n_particles)
    for calc in range(n_particles):
        x_1 = X[0, calc]
        x_2 = X[1, calc]
        x_3 = X[2, calc]
        x_4 = X[3, calc]
        x_5 = X[4, calc]
        x_6 = X2[0, calc]
        x_7 = X2[1, calc]
        x_8 = X2[2, calc]
        if x_1 < 0:
            x_1 = X[0, calc] = abs(x_1)
        if x_2 < 0:
            x_2 = X[1, calc] = abs(x_2)
        if x_3 < 0:
            x_3 = X[2, calc] = abs(x_3)
        if x_4 < 0:
            x_4 = X[0, calc] = abs(x_4)
        if x_5 < 0:
            x_5 = X[1, calc] = abs(x_5)
        if x_6 < 0:
            x_6 = X2[0, calc] = abs(x_6)
        if x_7 < 0:
            x_7 = X2[1, calc] = abs(x_7)
        if x_8 < 0:
            x_8 = X2[2, calc] = abs(x_8)
        Output = ObjectiveFunction(x_1, x_2, x_3, x_4, x_5, x_6, x_7, x_8)
        obj[calc] = Output[0]
        obj2[calc] = Output[1]
        with open(SystemPath + "\\Output\\ParaFitCryst_CalcuLog.txt", "a") as myfile:
            myfile.write("{}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}\n".format(SimuNo,
                                                                           datetime.now().strftime("%H:%M:%S"),
                                                                           "PSO",
                                                                           obj[calc],
                                                                           obj2[calc],
                                                                           x_1, x_2, x_3, x_4, x_5, x_6, x_7, x_8,  "PSO"))
    print([1, SimuNo,
        obj.min(), obj.mean(),
        X[0][obj.argmin()],
        X[1][obj.argmin()],
        X[2][obj.argmin()],
        X[3][obj.argmin()],
        X[4][obj.argmin()],
        obj2.min(), obj2.mean(),
        X2[0][obj2.argmin()],
        X2[1][obj2.argmin()], 
        X2[2][obj2.argmin()],   
        "PSO"])
    sys.stdout.flush()

    pbest[:, (pbest_obj >= obj)] = X[:, (pbest_obj >= obj)]
    pbest_obj = np.array([pbest_obj, obj]).min(axis=0)
    gbest = pbest[:, pbest_obj.argmin()]
    gbest_obj = pbest_obj.min()

    pbest2[:, (pbest_obj2 >= obj2)] = X2[:, (pbest_obj2 >= obj2)]
    pbest_obj2 = np.array([pbest_obj2, obj2]).min(axis=0)
    gbest2 = pbest2[:, pbest_obj2.argmin()]
    gbest_obj2 = pbest_obj2.min()



c1 = c2 = 0.5
w = 0.5
initial_particles = 50
n_particles = initial_particles

##### Currently, two approaches for the crystal growth rate are employed. Depending if the Exp. approach or the BCF approach is used different
###### values are used to span the starting space.
####### If the BCF approach is used, the third parameter serves no purpose. It is calculated nevertheless to simplifiy subsequent parts of the script
if GrowthRateApproach == "Exponential":
    x_ranges = [(500000, 1000000),(50000, 75000),(1, 2), (20, 250),(1, 10)]
else:
    x_ranges = [(0.1, 100),(0.01, 0.5),(1, 2),(20, 250),(50, 1000)]
x_ranges2 = [(0.5, 10), (0.01, 1), (1, 5)]
x_ranges3 = [(20, 250),(1, 10)]

X = np.array([np.linspace(x_range[0], x_range[1], n_particles) for x_range in x_ranges])
X2 = np.array([np.linspace(x_range[0], x_range[1], n_particles) for x_range in x_ranges2])
#X3 = np.array([np.random.uniform(x_range[0], x_range[1], n_particles) for x_range in x_ranges3])
V = np.random.randn(5, n_particles) * 0.1
V2 = np.random.randn(3, n_particles) * 0.1
#V3 = np.random.randn(2, n_particles) * 0.1
pbest = X
pbest2 = X2
#pbest3 = X3
pbest_obj = np.empty(n_particles)
pbest_obj2 = np.empty(n_particles)
#pbest_obj3 = np.empty(n_particles)
for calc in range(n_particles):
    x_1 = X[0, calc]
    x_2 = X[1, calc]
    x_3 = X[2, calc]
    x_4 = X[3, calc]
    x_5 = X[4, calc]
    x_6 = X2[0, calc]
    x_7 = X2[1, calc]
    x_8 = X2[2, calc]
    if x_1 < 0:
        x_1 = X[0, calc] = abs(x_1)
    if x_2 < 0:
        x_2 = X[1, calc] = abs(x_2)
    if x_3 < 0:
        x_3 = X[2, calc] = abs(x_3)
    if x_4 < 0:
        x_4 = X[0, calc] = abs(x_4)
    if x_5 < 0:
        x_5 = X[1, calc] = abs(x_5)
    if x_6 < 0:
        x_6 = X2[0, calc] = abs(x_6)
    if x_7 < 0:
        x_7 = X2[1, calc] = abs(x_7)
    if x_8 < 0:
        x_8 = X2[2, calc] = abs(x_8)
    Output = ObjectiveFunction(x_1, x_2, x_3, x_4, x_5, x_6, x_7, x_8)
    obj_result = Output[0]
    obj_result2 = Output[1]
    #obj_result3 = Output[2]
    pbest_obj[calc] = obj_result
    pbest_obj2[calc] = obj_result2
    #pbest_obj3[calc] = obj_result3
    with open(SystemPath + "\\Output\\ParaFitCryst_CalcuLog.txt", "a") as myfile:
        myfile.write("{}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}\n".format(0,
                                                                        datetime.now().strftime("%H:%M:%S"),
                                                                        "Initial Swarm",
                                                                        pbest_obj[calc],
                                                                        pbest_obj2[calc],
                                                                        x_1,
                                                                        x_2,
                                                                        x_3,
                                                                        x_4,
                                                                        x_5,
                                                                        x_6,
                                                                        x_7,
                                                                        x_8, "PSO"))

sorted_indices = np.argsort(pbest_obj)
selected_indices = sorted_indices[:initial_particles // 2]

NumSorted_indices = np.argsort(pbest_obj2)
NumSelected_indices = NumSorted_indices[:initial_particles // 2]

gbest = pbest[:, pbest_obj.argmin()]
gbest_obj = pbest_obj.min()
gbest2 = pbest2[:, pbest_obj2.argmin()]
gbest_obj2 = pbest_obj2.min()
#gbest3 = pbest3[:, pbest_obj3.argmin()]
#gbest_obj3 = pbest_obj3.min()

X = X[:, selected_indices]
X2 = X2[:, NumSelected_indices]
V = V[:, selected_indices]
V2 = V2[:, NumSelected_indices]
pbest = pbest[:, selected_indices]
pbest2 = pbest2[:, NumSelected_indices]
pbest_obj = pbest_obj[selected_indices]
pbest_obj2 = pbest_obj2[NumSelected_indices]
n_particles = initial_particles // 2

for it in range(1, Iterations + 1):
    update(it)

    if it == Iterations // 4:
        sorted_indices2 = np.argsort(pbest_obj)
        selected_indices2 = sorted_indices2[:n_particles // 2]

        NumSorted_indices2 = np.argsort(pbest_obj2)
        NumSelected_indices2 = NumSorted_indices2[:n_particles //2]

        X = X[:, selected_indices2]
        X2 = X2[:, NumSelected_indices2]
        V = V[:, selected_indices2]
        V2 = V2[:, NumSelected_indices2]
        pbest = pbest[:, selected_indices2]
        pbest2 = pbest2[:, NumSelected_indices2]
        pbest_obj = pbest_obj[selected_indices2]
        pbest_obj2 = pbest_obj2[NumSelected_indices2]
        n_particles = n_particles // 2

    print([1, it,
            gbest_obj,
            mean(pbest_obj),
            gbest[0],
            gbest[1],
            gbest[2],
            gbest[3],
            gbest[4],
            gbest_obj2,
            gbest2[0],
            gbest2[1],
            gbest2[2], "PSO"])
    with open(SystemPath + "\\Output\\ParaFitCryst_ObjectiveLog.txt", "a") as myfile:
        myfile.write("{}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}\n".format(it,
                                                                    datetime.now().strftime("%H:%M:%S"),
                                                                    gbest_obj,
                                                                    pbest_obj.mean(),
                                                                    gbest_obj2,
                                                                    gbest[0],
                                                                    gbest[1],
                                                                    gbest[2],
                                                                    gbest[3],
                                                                    gbest[4],
                                                                    gbest2[0],
                                                                    gbest2[1],
                                                                    gbest2[2], "PSO"))        


print([0, it,
            gbest_obj,
            pbest_obj.mean(),
            gbest[0],
            gbest[1],
            gbest[2],
            gbest[3],
            gbest[4],
            gbest_obj2,
            gbest2[0],
            gbest2[1],
            gbest2[2], "PSO"])