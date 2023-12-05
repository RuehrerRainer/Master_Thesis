######## Combined script for Particle Swarm Optimization and Nelder-Mead Algorithm ########

import numpy as np
import openpyxl
import sys
import os
import subprocess
import json
import time
from datetime import datetime
from statistics import mean
import random

SystemPath = os.path.dirname(__file__)

#ExDataPath = str(sys.argv[1])
ExDataPath = "C:\\Users\\Maxim\\Desktop\\Input\\ExperimentalData_Prog2.xlsx"
#Iterations = int(float(sys.argv[2]))
Iterations = 50
#NelderMead, PSO = str(sys.argv[3]), str(sys.argv[4])
NelderMead = False 
PSO = "True"

startTime = datetime.now()

#Prepare log file
with open(SystemPath + "\\DWSIM\\Output\\ParaFitCryst_CalcuLog.txt", "w") as myfile:
    myfile.write("Simulation number, Time, Type, Objective function, Growth Parameter 1, Growth Parameter 2, Growth Parameter 3, Agglomeration Parameter 1, Agglomeration Parameter 2, Fitting algorithm \n")
with open(SystemPath + "\\DWSIM\\Output\\ParaFitCryst_ObjectiveLog.txt", "w") as myfile:
    myfile.write("Simulation number, Time, Objective function, Mean Objective, Growth Parameter 1, Growth Parameter 2, Growth Parameter 3, Agglomeration Parameter 1, Agglomeration Parameter 2, Fitting algorithm\n")

#Read Experimental Data from input workbook
#The input workbook is passed as input argument by the subprocess
ExData = openpyxl.load_workbook(ExDataPath)
ExDataSheet = ExData['Results']
Ex10, Ex25, Ex50, Ex75, Ex90 = np.zeros(4), np.zeros(4), np.zeros(4), np.zeros(4), np.zeros(4)

for i in range(4):
    Ex10[i] = ExDataSheet.cell(row=i+2, column=2).value
    Ex25[i] = ExDataSheet.cell(row=i+2, column=3).value
    Ex50[i] = ExDataSheet.cell(row=i+2, column=4).value
    Ex75[i] = ExDataSheet.cell(row=i+2, column=5).value
    Ex90[i] = ExDataSheet.cell(row=i+2, column=6).value
ExData.close()

#SimResults stores the calculated diameters from the simulation for the respective time
SimResults10, SimResults25, SimResults50, SimResults75, SimResults90 = np.zeros(4), np.zeros(4), np.zeros(4), np.zeros(4), np.zeros(4)
#CSD is used to store the particle size distributions obtained in the simulations
CSD30, CSD60, CSD90, CSD120 = np.zeros(30), np.zeros(30), np.zeros(30), np.zeros(30)

def Load_JSON_file(filepath):
    """
    Loads the .json file used for storing the input parameters in the DWSIM simulation.
    The data is stored as dictionary and returned by this function
    """
    JSON_file = open(filepath + "\\DWSIM\\Input.json")
    JSON_data = json.load(JSON_file)
    return JSON_data

#Initialize parameters used in the calculation
Parameters = Load_JSON_file(SystemPath)
Time = int(int(Parameters["CycleTime"]) * int(Parameters["CrystallizationModules"]))
Toleranz = 10
delta = int(Parameters["Delta"])

GrowthRateApproach = Parameters["GrowthRate"]

#Initialize lists used in the calculation
PenaltySimp = np.zeros(6) #Used to store values of the objective function for the simplex
ResCSD = np.zeros(30)


#Import from the opened JSON file
Widths = Parameters["ClassWidth"]
Length = Parameters["Length"]

def ImportSimulationOutput():
    """
    Loads the results of the simulation, converts it from a density function to a sum function and calculates dxx.
    The parameters are returned in a list for further use in calculations
    """

    SimOutput = openpyxl.load_workbook(SystemPath + '\\DWSIM\\Output\\CrystallizationOutput.xlsx')
    SimOut = SimOutput['Results']
    for Ex in range(30):
        CSD30[Ex] = SimOut.cell(row=Ex+9, column=(1+(30*60)/delta)).value
        CSD60[Ex] = SimOut.cell(row=Ex+9, column=(1+(60*60)/delta)).value
        CSD90[Ex] = SimOut.cell(row=Ex+9, column=(1+(90*60)/delta)).value
        CSD120[Ex] = SimOut.cell(row=Ex+9, column=(1+(120*60)/delta)).value
    Distros = [CSD30, CSD60, CSD90, CSD120]

    #Auxiliary lists for the calculation:
    NumDest, Q3 = np.zeros(30), np.zeros(30)

    for j in range(4):
        NumDest_tot = 0
        # for i in range(0, 30):
        #     #Conversion of the density function to a cumulative function
        #     ResCSD[i] = Distros[j][i] * Widths[i]
        #     if i == 0:
        #         ResCSD[i] = ResCSD[i]
        #     else:
        #         ResCSD[i] = ResCSD[i] + ResCSD[i-1]
        # for i in range(0, 30):
        #     #Normalization of the distribution function
        #     #Done to mitigate numerical instabilities and numerical diffusion
        #     ResCSD[i] = ResCSD[i] / ResCSD[29]

        #     if i == 0:
        #         NumDest[i] = ResCSD[i]
        #     else:
        #         NumDest[i] = (ResCSD[i] - ResCSD[i-1])
        #     NumDest_i = NumDest[i] * Length[i]**3
        #     NumDest_tot = NumDest_tot + NumDest_i
        # for i in range(0, 30):
        #     #Conversion of the number density distribution to a Q3 distribution
        #     Q3[i] = NumDest[i] * Length[i]**3 / NumDest_tot
        #     if i == 0:
        #         Q3[i] = Q3[i]
        #     else:
        #         Q3[i] = Q3[i] + Q3[i-1]

        for i in range(0, 30):
            if i == 0:
                Q3[i] = Distros[j][i]
            else:
                Q3[i] = Distros[j][i] + Q3[i-1]
        for i in range(1, 30):
            #
            if Q3[i] < 0.01:
                i+=1
            #if (Q3[i] >= 0.01 and Q3[i-1] < 0.01).all():
            if np.all(np.logical_and(Q3[i] >= 0.10, Q3[i-1] < 0.10)):
                SimResults10[j] = (Length[i] - Length[i-1])/(Q3[i] - Q3[i-1]) * (0.10 - Q3[i-1]) + Length[i-1]
            elif np.all(np.logical_and(Q3[i] >= 0.25, Q3[i-1] < 0.25)):
                SimResults25[j] = (Length[i] - Length[i-1])/(Q3[i] - Q3[i-1]) * (0.25 - Q3[i-1]) + Length[i-1]
            elif np.all(np.logical_and(Q3[i] >= 0.50, Q3[i-1] < 0.50)):
                SimResults50[j] = (Length[i] - Length[i-1])/(Q3[i] - Q3[i-1]) * (0.50 - Q3[i-1]) + Length[i-1]
            elif np.all(np.logical_and(Q3[i] >= 0.75, Q3[i-1] < 0.75)):
                SimResults75[j] = (Length[i] - Length[i-1])/(Q3[i] - Q3[i-1]) * (0.75 - Q3[i-1]) + Length[i-1]
            elif np.all(np.logical_and(Q3[i] >= 0.90, Q3[i-1] < 0.90)):
                SimResults90[j] = (Length[i] - Length[i-1])/(Q3[i] - Q3[i-1]) * (0.90 - Q3[i-1]) + Length[i-1]
                break
    SimOutput.close()
    os.remove(SystemPath + '\\DWSIM\\Output\\CrystallizationOutput.xlsx')
    return [SimResults10, SimResults25, SimResults50, SimResults75, SimResults90]

def PenaltyCalculation(lstSimulationOutput):
    """
    Calculation of the penalty function which is minimized over the course of this script.
    The input for this function is a list containing lists containing the time dependent results for
    d01, d10, d50, d90 and d99
    """
    #Objective_i_10 = ((lstSimulationOutput[0][0] - Ex10[0])**2 + (lstSimulationOutput[0][1]-Ex10[1])**2 + (lstSimulationOutput[0][2]-Ex10[2])**2 +(lstSimulationOutput[0][3]-Ex10[3])**2)
    #Objective_i_25 = ((lstSimulationOutput[1][0] - Ex25[0])**2 + (lstSimulationOutput[1][1]-Ex25[1])**2 + (lstSimulationOutput[1][2]-Ex25[2])**2 +(lstSimulationOutput[1][3]-Ex25[3])**2)
    #Objective_i_50 = ((lstSimulationOutput[2][0] - Ex50[0])**2 + (lstSimulationOutput[2][1]-Ex50[1])**2 + (lstSimulationOutput[2][2]-Ex50[2])**2 +(lstSimulationOutput[2][3]-Ex50[3])**2)
    #Objective_i_75 = ((lstSimulationOutput[3][0] - Ex75[0])**2 + (lstSimulationOutput[3][1]-Ex75[1])**2 + (lstSimulationOutput[3][2]-Ex75[2])**2 +(lstSimulationOutput[3][3]-Ex75[3])**2)
    #Objective_i_90 = ((lstSimulationOutput[4][0] - Ex90[0])**2 + (lstSimulationOutput[4][1]-Ex90[1])**2 + (lstSimulationOutput[4][2]-Ex90[2])**2 +(lstSimulationOutput[4][3]-Ex90[3])**2)
    #Objective_i_Width = ((lstSimulationOutput[4][0] - lstSimulationOutput[0][0]) - (Ex90[0] - Ex10[0]))**2 + ((lstSimulationOutput[4][1] - lstSimulationOutput[0][1]) - (Ex90[1] - Ex10[1]))**2 + ((lstSimulationOutput[4][2] - lstSimulationOutput[0][2]) - (Ex90[2] - Ex10[2]))**2 + ((lstSimulationOutput[4][3] - lstSimulationOutput[0][3]) - (Ex90[3] - Ex10[3]))**2
    
    Objective_i_10 = (lstSimulationOutput[0][3]-Ex10[3])**2
    Objective_i_25 = (lstSimulationOutput[1][3]-Ex25[3])**2
    Objective_i_50 = (lstSimulationOutput[2][3]-Ex50[3])**2
    Objective_i_75 = (lstSimulationOutput[3][3]-Ex75[3])**2
    Objective_i_90 = (lstSimulationOutput[4][3]-Ex90[3])**2
    Objective_i_Width = ((lstSimulationOutput[4][3] - lstSimulationOutput[0][3]) - (Ex90[3] - Ex10[3]))**2

    #Objective = Objective_i_10 + Objective_i_25 + Objective_i_50 + Objective_i_75 + Objective_i_90 + Objective_i_Width
    Objective = Objective_i_50
    return np.sqrt(Objective)
    
def Calculate():
    """
    This function calls the DWSIM solver using the "subprocess.call" method.
    The solver is called via an additional script to prevent a synchrounous calculation of this script and the DWSIM simulation which would cause erroneous results.
    """
    Process = subprocess.Popen(['python', SystemPath + '\\DWSIM\\RunCrystallization.py', SystemPath + "\\DWSIM\\DWSIM_Files\\", "Fitting"], text=True)
    Process.wait()

def GiveInputs(lstIn):
    """
    Passes arguments for growth and agglomeration parameters to the .JSON file which is subsequently read by the DWSIM script.
    """
    JSONFile = Load_JSON_file(SystemPath)
    JSONFile["GrowthConstant1"] = lstIn[0]
    JSONFile["GrowthConstant2"] = lstIn[1]
    JSONFile["GrowthConstant3"] = lstIn[2]
    JSONFile["AgglConstant1"] = lstIn[3]
    JSONFile["AgglConstant2"] = lstIn[4]
    JSONFile["BirthConstant1"] = lstIn[5]
    JSONFile["BirthConstant2"] = lstIn[6]
    JSONFile["BirthConstant3"] = lstIn[7]
    with open(SystemPath + "\\Input.json", "w") as json_file:
        json.dump(JSONFile, json_file, indent=4)

####################### Functions specific for PSO #######################
def ObjectiveFunction(para1:float, para2:float, para3:float, para4:float, para5:float, para6:float, para7:float, para8:float):

    GiveInputs([para1, para2, para3, para4, para5, para6, para7, para8])
    Calculate()
    Output = PenaltyCalculation(ImportSimulationOutput())
    return Output
    
def update(SimuNo:int):
    global V, X, pbest, pbest_obj, gbest, gbest_obj

    r1, r2 = np.random.rand(2)
    V = w * V + c1 * r1 *(pbest - X) + c2 * r2 * (gbest.reshape(-1,1) - X)
    X = X + V
    
    obj = np.empty(n_particles)
    for calc in range(n_particles):
        x_1 = X[0, calc]
        x_2 = X[1, calc]
        x_3 = X[2, calc]
        x_4 = X[3, calc]
        x_5 = X[4, calc]
        x_6 = X[5, calc]
        x_7 = X[6, calc]
        x_8 = X[7, calc]
        if x_5 <= 0:
            x_5 = 1e-9
        obj[calc] = ObjectiveFunction(x_1, x_2, x_3, x_4, x_5, x_6, x_7, x_8)
        with open(SystemPath + "\\DWSIM\\Output\\ParaFitCryst_CalcuLog.txt", "a") as myfile:
            myfile.write("{}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}\n".format(SimuNo,
                                                                           datetime.now().strftime("%H:%M:%S"),
                                                                           "PSO",
                                                                           obj[calc],
                                                                           x_1, x_2, x_3, x_4, x_5, x_6, x_7, x_8,  "PSO"))
        print([1, SimuNo, min(obj), mean(obj),
            X[0][obj.argmin()],
            X[1][obj.argmin()],
            X[2][obj.argmin()],
            X[3][obj.argmin()],
            X[4][obj.argmin()],
            X[5][obj.argmin()],
            X[6][obj.argmin()], 
            X[7][obj.argmin()],   
            "PSO"])
        sys.stdout.flush()

    pbest[:, (pbest_obj >= obj)] = X[:, (pbest_obj >= obj)]
    pbest_obj = np.array([pbest_obj, obj]).min(axis=0)
    gbest = pbest[:, pbest_obj.argmin()]
    gbest_obj = pbest_obj.min()
    
####################### Functions specific for Nelder Mead #######################
def PopulateInitSimplex():
    global Simplex

    if GrowthRateApproach == "Exponential":
        for i in range(1,7):
            for i in range(1,7):
                Simplex[i][0] = random.uniform(10, 1000) #Prefactor
                Simplex[i][1] = random.uniform(1000, 100000) #Activation energy
                Simplex[i][2] = random.uniform(1, 2) #Exponential
                Simplex[i][3] = random.uniform(20, 250) #Critical length
                Simplex[i][4] = random.uniform(0.00000001, 0.00001) #Agglomeration parameter
                Simplex[i][5] = random.uniform(0.01, 5)
                Simplex[i][6] = random.uniform(0, 2)
                Simplex[i][7] = random.uniform(1, 5)
    else:
        for i in range(1,7):
            Simplex[i][0] = random.uniform(0.1, 1000) #A_BCF
            Simplex[i][1] = random.uniform(0, 5) #B_BCF
            Simplex[i][2] = random.uniform(0, 0) #Unused
            Simplex[i][3] = random.uniform(20, 250) #Critical length
            Simplex[i][4] = random.uniform(0.00000001, 0.01) #Agglomeration parameter
            Simplex[i][5] = random.uniform(0.1, 10000)
            Simplex[i][6] = random.uniform(0, 2)
            Simplex[i][7] = random.uniform(1, 5)

def AverageSimplex():
    """
    Averages the input values stored in the simplex dictionary.
    Used in the calculation of the reflection, expansion and contraction step.
    """
    Ave = np.empty(8)
    for value in range(8):
        Ave[value] = (Simplex[1][value] + Simplex[2][value] + Simplex[3][value] + Simplex[4][value] + Simplex[5][value] + Simplex[6][value] + Simplex[7][value] + Simplex[8][value])/6
    return Ave

def ReplaceSimplex(arrVals, SimpNo):
    """
    This function is called if either the reflection, expansion or contraction step prove to be successful.
    By replacing the values in the simplex, the original one can be kept throughout the script instead of being calculated again for each iteration.
    """
    for Repl in range(8):
        Simplex[SimpNo][Repl] = arrVals[Repl]

def ReplaceMin(fltVal, MinNo):
    PenaltySimplex[MinNo] = fltVal

def Reflection(lstAve, intNumber):
    """
    Application of the reflection step in the Nelder-mead method.
    Initial step performed after the initial simplex calculation.
    """    
    alpha = 1 #Parameter defining the step width in the reflection calculation
    Reflect = abs((1 + alpha) * lstAve - alpha * Simplex[list(PenaltySimplex).index(max(PenaltySimplex)) + 1])
    GiveInputs(Reflect)
    Calculate()
    PenaltyReflection = PenaltyCalculation(ImportSimulationOutput())
    #print(PenaltyReflection)
    with open(SystemPath + "\\DWSIM\\Output\\ParaFitCryst_CalcuLog.txt", "a") as myfile:
        myfile.write("{}, {}, {}, {}, {}, {}, {}, {}, {}, {}\n".format(intNumber,
                                                                       datetime.now().strftime("%H:%M:%S"),
                                                                       "Reflection",
                                                                       PenaltyReflection,
                                                                        Reflect[0],
                                                                        Reflect[1],
                                                                        Reflect[2],
                                                                        Reflect[3],
                                                                        Reflect[4],
                                                                        Reflect[5],
                                                                        Reflect[6],
                                                                        Reflect[7]))
    return PenaltyReflection, Reflect

def Expansion(ReflecPoint, lstAve, intNumber):
    """
    Performs the expansion step in the Nelder-Mead method.
    The absolute value is used in the determination of input parameters in order to prevent negative values.
    Should negative values occur, the simulation in DWSIM cannot be performed accordingly.
    """
    gamma = 0.5 #Parameter used to define the extension of the simplex in the expansion step
    Expand = abs((1 + gamma) * ReflecPoint - gamma * lstAve)
    GiveInputs(Expand)
    Calculate()
    PenaltyExpansion = PenaltyCalculation(ImportSimulationOutput())
    #print(PenaltyExpansion)
    with open(SystemPath + "\\DWSIM\\Output\\ParaFitCryst_CalcuLog.txt", "a") as myfile:
        myfile.write("{}, {}, {}, {}, {}, {}, {}, {}, {}, {}\n".format(intNumber,
                                                                       datetime.now().strftime("%H:%M:%S"),
                                                                       "Expansion",
                                                                       PenaltyExpansion,
                                                                        Expand[0],
                                                                        Expand[1],
                                                                        Expand[2],
                                                                        Expand[3],
                                                                        Expand[4],
                                                                        Expand[5],
                                                                        Expand[6],
                                                                        Expand[7]))

def Contraction(lstAve, intNumber):
    """
    Performs the contraction step in the Nelder Mead method.
    """
    beta = 0.5 #Parameter defining the extent of the contraction
    Contract = abs(beta * Simplex[list(PenaltySimplex).index(min(PenaltySimplex)) + 1] + (1 - beta) * lstAve)
    GiveInputs(Contract)
    Calculate()
    PenaltyContraction = PenaltyCalculation(ImportSimulationOutput())
    #print(PenaltyContraction)
    with open(SystemPath + "\\DWSIM\\Output\\ParaFitCryst_CalcuLog.txt", "a") as myfile:
        myfile.write("{}, {}, {}, {}, {}, {}, {}, {}, {}, {}\n".format(intNumber,
                                                                       datetime.now().strftime("%H:%M:%S"),
                                                                       "Contraction",
                                                                        PenaltyContraction,
                                                                        Contract[0],
                                                                        Contract[1],
                                                                        Contract[2],
                                                                        Contract[3],
                                                                        Contract[4],
                                                                        Contract[5],
                                                                        Contract[6],
                                                                        Contract[7]))
        return PenaltyContraction, Contract

def NelderMeadMainLoop(SimulationNo:int):
    Average = AverageSimplex()
    ReflectionResults = Reflection(Average, SimulationNo)

    if ReflectionResults[0] <= Toleranz:
        #print("Final simplex:", Simplex[list(PenaltySimplex).index(min(PenaltySimplex)) + 1])
        #print("Final value of optimization function:", min(PenaltySimplex))
        print([1, SimulationNo,
               min(PenaltySimplex),
               mean(PenaltySimplex),
               Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][0],
               Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][1],
               Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][2],
               Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][3],
               Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][4],
                Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][5],
                Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][6],
                Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][7], "NM"])
        sys.stdout.flush()
        sys.exit()
    else: 
        if ReflectionResults[0] < min(PenaltySimplex):
            ExpansionResults = Expansion(ReflectionResults[1], Average, SimulationNo)

            if ExpansionResults[0] <= Toleranz:
                #print("Optimal value:", ExpansionResults[1])
                #print("Final value of optimization function:", ExpansionResults[0])
                print([1, SimulationNo,
                       min(PenaltySimplex),
                       mean(PenaltySimplex),
                       Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][0],
                       Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][1],
                       Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][2],
                       Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][3],
                       Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][4],
                        Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][5],
                        Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][6],
                        Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][7], "NM"])
                sys.stdout.flush()
                sys.exit()

            else:
                if ExpansionResults[0] < min(PenaltySimplex):
                    ReplaceSimplex(ExpansionResults[1], list(PenaltySimplex).index(max(PenaltySimplex))+1)
                    ReplaceMin(ExpansionResults[0], list(PenaltySimplex).index(max(PenaltySimplex)))
                    #print(PenaltySimplex)
                else:
                    ReplaceSimplex(ReflectionResults[1], list(PenaltySimplex).index(max(PenaltySimplex))+1)
                    ReplaceMin(ReflectionResults[0], list(PenaltySimplex).index(max(PenaltySimplex)))
                    #print(PenaltySimplex)
        ##### Middle branch of the Nelder Mead algorithm #####
        elif any(i > ReflectionResults[0] for i in PenaltySimplex if i != max(PenaltySimplex)):
            ReplaceSimplex(ReflectionResults[1], list(PenaltySimplex).index(max(PenaltySimplex))+1)
            ReplaceMin(ReflectionResults[0], list(PenaltySimplex).index(max(PenaltySimplex)))
            #print(PenaltySimplex)
        
        ##### Right branch of the Nelder Mead algorithm #####
        else:

            if ReflectionResults[0] < max(PenaltySimplex):
                ReplaceSimplex(ReflectionResults[1], list(PenaltySimplex).index(max(PenaltySimplex))+1)
                ReplaceMin(ReflectionResults[0], list(PenaltySimplex).index(max(PenaltySimplex)))
                Average = AverageSimplex()
                #print(PenaltySimplex)
            
            ##### Contraction step #####
            ContractionResults = Contraction(Average, SimulationNo)

            if ContractionResults[0] <= Toleranz:
                #print("Optimal value:", ContractionResults[1])
                #print("Final value of optimization function:", ContractionResults[0])
                print([1, SimulationNo,
                       min(PenaltySimplex),
                       mean(PenaltySimplex),
                       Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][0],
                       Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][1],
                       Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][2],
                       Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][3],
                       Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][4],
                        Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][5],
                        Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][6],
                        Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][7], "NM"])
                sys.stdout.flush()
                sys.exit()

            elif ContractionResults[0] < max(PenaltySimplex):
                ReplaceSimplex(ContractionResults[1], list(PenaltySimplex).index(max(PenaltySimplex))+1)
                ReplaceMin(ContractionResults[0], list(PenaltySimplex).index(max(PenaltySimplex)))
                #print(PenaltySimplex)

            else:
                for i in range(1,7):
                    Simplex[i] = abs((Simplex[i]+ Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1])/2)
                Average = AverageSimplex()
                #print("Calculating compressed simplex...")
                for i in range(1,7):
                    GiveInputs(Simplex[i])
                    Calculate()
                    PenaltySimplex[i-1] = PenaltyCalculation(ImportSimulationOutput())
                    with open(SystemPath + "\\DWSIM\\Output\\ParaFitCryst_CalcuLog.txt", "a") as myfile:
                        myfile.write("{}, {}, {}, {}, {}, {}, {}, {}, {}, {}\n".format(SimulationNo,
                                                                                       datetime.now().strftime("%H:%M:%S"),
                                                                                       "Compression",
                                                                                       PenaltySimplex[i-1],
                                                                                       Simplex[i][0],
                                                                                       Simplex[i][1],
                                                                                       Simplex[i][2],
                                                                                       Simplex[i][3],
                                                                                       Simplex[i][4], "NM"))
                #print("Finished compressed simplex calculation.")
                #print(PenaltySimplex)
                if min(PenaltySimplex) <= Toleranz:
                    #print("Optimal value:", Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1])
                    #print("Final value of optimization function: ", min(PenaltySimplex))
                    print([1, SimulationNo,
                           min(PenaltySimplex),
                           mean(PenaltySimplex),
                           Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][0],
                           Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][1],
                           Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][2],
                           Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][3],
                           Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][4],
                            Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][5],
                            Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][6],
                            Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][7], "NM"])
                    sys.stdout.flush()
                    sys.exit()
    #print("Minimum of the objective function in iteration %i was:" %SimulationNo, min(PenaltySimplex))
    with open(SystemPath + "\\DWSIM\\Output\\ParaFitCryst_ObjectiveLog.txt", "a") as myfile:
        myfile.write("{}, {}, {}, {}, {}, {}, {}, {}, {}, {}\n".format(SimulationNo,
                                                                       datetime.now().strftime("%H:%M:%S"),
                                                                       min(PenaltySimplex),
                                                                       mean(PenaltySimplex),
                                                                       Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][0],
                                                                       Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][1],
                                                                       Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][2],
                                                                       Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][3],
                                                                       Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][4],
                                                                        Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][5],
                                                                        Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][6],
                                                                        Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][7], "NM"))
    if SimulationNo != Iterations:
        print([0, SimulationNo,
               min(PenaltySimplex),
               mean(PenaltySimplex),
               Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][0],
               Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][1],
               Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][2],
               Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][3],
               Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][4],
                Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][5],
                Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][6],
                Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][7], "NM"])
        sys.stdout.flush()

##### Initialization of the simplex used for the Nelder Mead Algorithm and its counterpart used to store the corresponding values of the objective function
Simplex = {
    1: np.zeros(8),
    2: np.zeros(8),
    3: np.zeros(8),
    4: np.zeros(8),
    5: np.zeros(8),
    6: np.zeros(8)
}
PenaltySimplex = [None] * 6

##### Main loop: In the case that PSO and NM are set to be active, PSO is used to determine the starting points for the NM algorithm
########## If NM is not active, PSO is used to calculate the optimum
if PSO == "True":
    c1 = c2 = 2
    w = 0.8
    n_particles = 25

    ##### Currently, two approaches for the crystal growth rate are employed. Depending if the Exp. approach or the BCF approach is used different
    ###### values are used to span the starting space.
    ####### If the BCF approach is used, the third parameter serves no purpose. It is calculated nevertheless to simplifiy subsequent parts of the script
    if GrowthRateApproach == "Exponential":
        x_ranges = [(5, 100),(1000, 100000),(1, 2),(20, 250),(0.00000001, 0.01), (0.01, 5), (0, 2), (1, 5)]
    else:
        x_ranges = [(0, 1000),(0, 5),(1, 2),(20, 250),(0.00000001, 0.01), (0.01, 5), (0, 2), (1, 5)]

    X = np.array([np.random.uniform(x_range[0], x_range[1], n_particles) for x_range in x_ranges])
    V = np.random.randn(5, n_particles) * 0.1
    pbest = X
    pbest_obj = np.empty(n_particles)
    for calc in range(n_particles):
        x_1 = X[0, calc]
        x_2 = X[1, calc]
        x_3 = X[2, calc]
        x_4 = X[3, calc]
        x_5 = X[4, calc]
        x_6 = X[5, calc]
        x_7 = X[6, calc]
        x_8 = X[7, calc]
        if x_5 <= 0:
            x_5 = 1e-9
        obj_result = ObjectiveFunction(x_1, x_2, x_3, x_4, x_5, x_6, x_7, x_8)
        pbest_obj[calc] = obj_result
        with open(SystemPath + "\\DWSIM\\Output\\ParaFitCryst_CalcuLog.txt", "a") as myfile:
            myfile.write("{}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}\n".format(0,
                                                                           datetime.now().strftime("%H:%M:%S"),
                                                                           "Initial Swarm",
                                                                           pbest_obj[calc],
                                                                           x_1,
                                                                           x_2,
                                                                           x_3,
                                                                           x_4,
                                                                           x_5,
                                                                           x_6,
                                                                           x_7,
                                                                           x_8, "PSO"))

    gbest = pbest[:, pbest_obj.argmin()]
    gbest_obj = pbest_obj.min()

    for it in range(1, Iterations + 1):
        update(it)
        print([1, it,
               gbest_obj,
               mean(pbest_obj),
               gbest[0],
               gbest[1],
               gbest[2],
               gbest[3],
               gbest[4],
                gbest[5],
                gbest[6],
                gbest[7], "PSO"])
        with open(SystemPath + "\\DWSIM\\Output\\ParaFitCryst_ObjectiveLog.txt", "a") as myfile:
            myfile.write("{}, {}, {}, {}, {}, {}, {}, {}, {}, {}\n".format(it,
                                                                       datetime.now().strftime("%H:%M:%S"),
                                                                       gbest_obj,
                                                                       mean(pbest_obj),
                                                                       gbest[0],
                                                                       gbest[1],
                                                                       gbest[2],
                                                                       gbest[3],
                                                                       gbest[4],
                                                                        gbest[5],
                                                                        gbest[6],
                                                                        gbest[7], "PSO"))        
    Bests = np.array(pbest_obj, pbest)
    sorted_indices = np.argsort(Bests[0])
    sorted_pbest_obj = Bests[0][sorted_indices]
    sorted_pbest = Bests[1][:, sorted_indices]
    Bests = [sorted_pbest_obj, sorted_pbest]

    print([0, it,
               gbest_obj,
               mean(pbest_obj),
               gbest[0],
               gbest[1],
               gbest[2],
               gbest[3],
               gbest[4],
                gbest[5],
                gbest[6],
                 gbest[7], "PSO"])


    if NelderMead == "True":
        for i in range(6):
            Simplex[i + 1][0] = Bests[1][0][i]
            Simplex[i + 1][1] = Bests[1][1][i]
            Simplex[i + 1][2] = Bests[1][2][i]
            Simplex[i + 1][3] = Bests[1][3][i]
            Simplex[i + 1][4] = Bests[1][4][i]
            Simplex[i + 1][5] = Bests[1][5][i]
            Simplex[i + 1][6] = Bests[1][6][i]
            Simplex[i + 1][7] = Bests[1][7][i]
            PenaltySimplex[i] = Bests[0][i]
            print(Simplex(i + 1))

        for it2 in range(1, Iterations + 1):
            NelderMeadMainLoop(it2 + it)
            print([1, it2 + it,
                   min(PenaltySimplex),
                   mean(PenaltySimplex),
                   Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][0],
                   Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][1],
                   Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][2],
                   Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][3],
                   Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][4],
                    Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][5],
                     Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][6],
                      Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][7], "NM"])
            sys.stdout.flush()

    
if NelderMead == "True" and PSO == "False":
    PopulateInitSimplex()
    SimulationNo = 1
    for i in range(0, 6):
        GiveInputs(Simplex[i + 1])
        Calculate()
        PenaltySimplex[i] = PenaltyCalculation(ImportSimulationOutput())
        #print(PenaltySimplex)
        with open(SystemPath + "\\DWSIM\\Output\\ParaFitCryst_CalcuLog.txt", "a") as myfile:
            myfile.write("{}, {}, {}, {}, {}, {}, {}, {}, {}, {}\n".format(SimulationNo,
                                                                           datetime.now().strftime("%H:%M:%S"),
                                                                           "Initial Simplex",
                                                                           PenaltySimplex[i],
                                                                           Simplex[i + 1][0],
                                                                           Simplex[i + 1][1],
                                                                           Simplex[i + 1][2],
                                                                           Simplex[i + 1][3],
                                                                           Simplex[i + 1][4],
                                                                           Simplex[i + 1][5],
                                                                           Simplex[i + 1][6],
                                                                           Simplex[i + 1][7],"NM"))

    if min(PenaltySimplex) <= Toleranz:
        print([1, SimulationNo,
               min(PenaltySimplex),
               mean(PenaltySimplex),
               Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][0],
               Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][1],
               Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][2],
               Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][3],
               Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][4],
                Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][5],
                 Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][6],
                  Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][7], "NM"])
        sys.stdout.flush()
        sys.exit()
    else:
        while SimulationNo <= Iterations:
            NelderMeadMainLoop(SimulationNo)
            print([1, SimulationNo,
                   min(PenaltySimplex),
                   mean(PenaltySimplex),
                   Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][0],
                   Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][1],
                   Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][2],
                   Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][3],
                   Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][4],
                    Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][5],
                     Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][6],
                      Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][7], "NM"])
            SimulationNo += 1

    print([1, SimulationNo,
           min(PenaltySimplex),
           mean(PenaltySimplex),
           Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][0],
           Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][1],
           Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][2],
           Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][3],
           Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][4],
            Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][5],
             Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][6],
              Simplex[list(PenaltySimplex).index(min(PenaltySimplex))+1][7], "NM"])
    sys.stdout.flush()    